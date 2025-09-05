from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook


class ExcelCalendarService:
    """Minimal Excel-based calendar with day view and 30/60-min slots.

    The workbook contains a single sheet with headers:
    [appointment_id, patient_id, doctor, location, start_iso, duration_minutes]
    """

    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        if not self.excel_path.exists():
            self._initialize()

    def _initialize(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "appointments"
        ws.append(["appointment_id", "patient_id", "doctor", "location", "start_iso", "duration_minutes"])
        wb.save(self.excel_path)

    def _load(self):
        return load_workbook(self.excel_path)

    def _list_bookings(self) -> List[dict]:
        wb = self._load()
        ws = wb["appointments"]
        bookings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            bookings.append(
                {
                    "appointment_id": row[0],
                    "patient_id": row[1],
                    "doctor": row[2],
                    "location": row[3],
                    "start_iso": row[4],
                    "duration_minutes": int(row[5]),
                }
            )
        return bookings

    def find_available_slots(
        self,
        preferred_doctor: Optional[str],
        preferred_location: Optional[str],
        duration_minutes: int,
        window_start: datetime,
        window_end: datetime,
        max_results: int = 5,
    ) -> List[datetime]:
        existing = self._list_bookings()
        busy = []
        for b in existing:
            start = datetime.fromisoformat(b["start_iso"]) if isinstance(b["start_iso"], str) else b["start_iso"]
            busy.append((start, timedelta(minutes=int(b["duration_minutes"]))))

        step = timedelta(minutes=30)
        slots: List[datetime] = []
        t = window_start
        while t + timedelta(minutes=duration_minutes) <= window_end and len(slots) < max_results:
            overlap = False
            for s, d in busy:
                if not (t + timedelta(minutes=duration_minutes) <= s or t >= s + d):
                    overlap = True
                    break
            if not overlap:
                slots.append(t)
            t += step
        return slots

    def book_slot(
        self,
        appointment_id: str,
        patient_id: str,
        doctor: str,
        location: str,
        start: datetime,
        duration_minutes: int,
    ) -> None:
        wb = self._load()
        ws = wb["appointments"]
        ws.append(
            [
                appointment_id,
                patient_id,
                doctor,
                location,
                start.isoformat(),
                int(duration_minutes),
            ]
        )
        wb.save(self.excel_path)
